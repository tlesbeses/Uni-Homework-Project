"""Excel export of section grades.

Builds an ``.xlsx`` workbook in memory with this layout:

    Curso:  <course title>
    Grupo:  <section name>
    ----------------------------------------------
    Estudiante | Assignment 1 | Assignment 2 | Total
    ...        | ...          | ...          | ...

Only published assignments and approved enrollments are included; missing
grades render as empty cells and ``Total`` sums the existing ones.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from assignments.models import Assignment
from course.models import Enrollment, Status
from grading.models import Grade

HEADER_ROW = 4


def _student_label(enrollment) -> str:
    student = enrollment.student
    return f"{student.first_name or student.username} {student.last_name or ''}".strip()


def build_section_grades_workbook(*, section) -> bytes:
    """Return the grades of ``section`` as an ``.xlsx`` byte string."""
    assignments = list(
        Assignment.objects.filter(
            course=section.course,
            is_published=True,
        ).order_by("due_date", "id")
    )

    enrollments = list(
        Enrollment.objects.filter(
            section=section,
            status=Status.APPROVED,
        )
        .select_related("student")
        .order_by("student__first_name", "student__last_name", "student__username")
    )

    enrollment_ids = [enrollment.student_id for enrollment in enrollments]
    assignment_ids = [assignment.id for assignment in assignments]
    scores_by_pair = {}
    if enrollment_ids and assignment_ids:
        for grade in Grade.objects.filter(
            assignment_id__in=assignment_ids,
            student_id__in=enrollment_ids,
        ).only("student_id", "assignment_id", "score"):
            scores_by_pair[(grade.student_id, grade.assignment_id)] = float(
                grade.score
            )

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Notas"
    bold = Font(bold=True)

    sheet["A1"] = "Curso:"
    sheet["B1"] = section.course.title
    sheet["A2"] = "Grupo:"
    sheet["B2"] = section.name
    sheet["A1"].font = bold
    sheet["A2"].font = bold

    headers = ["Estudiante", *[a.title for a in assignments], "Total"]
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(row=HEADER_ROW, column=column_index, value=header).font = bold

    total_column = len(assignments) + 2
    for offset, enrollment in enumerate(enrollments, start=1):
        row = HEADER_ROW + offset
        sheet.cell(row=row, column=1, value=_student_label(enrollment))
        total = 0.0
        for assignment_offset, assignment in enumerate(assignments, start=2):
            score = scores_by_pair.get((enrollment.student_id, assignment.id))
            if score is None:
                continue
            sheet.cell(row=row, column=assignment_offset, value=round(score, 2))
            total += score
        sheet.cell(row=row, column=total_column, value=round(total, 2))

    sheet.column_dimensions["A"].width = 28
    for column_index in range(2, total_column + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = 16

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()
