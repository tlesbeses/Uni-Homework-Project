"""Tests for the grading application.

Covers the 16 core business rules of the grading module.
"""

from decimal import Decimal
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.db import IntegrityError
from django.urls import reverse
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APITestCase

from assignments.models import Assignment
from authentication.models import EventLog
from course.models import Course, Enrollment, Section, Status
from grading.models import Grade
from teams.models import Team, TeamMember

User = get_user_model()


class GradingAPITestCase(APITestCase):
    """Shared fixtures for the grading API tests."""

    def setUp(self):
        self.teacher_group, _ = Group.objects.get_or_create(name="Teacher")

        self.teacher = self.create_user("teacher")
        self.teacher.groups.add(self.teacher_group)

        self.other_teacher = self.create_user("other_teacher")
        self.other_teacher.groups.add(self.teacher_group)

        self.student = self.create_user("student")
        self.student2 = self.create_user("student2")
        self.unapproved_student = self.create_user("unapproved_student")

        self.course = Course.objects.create(title="Math 101", teacher=self.teacher)
        self.other_course = Course.objects.create(
            title="Physics", teacher=self.other_teacher
        )

        self.enroll(self.student, self.course)
        self.enroll(self.student2, self.course)

        self.assignment = Assignment.objects.create(
            course=self.course,
            title="Homework 1",
            max_score="100.00",
            is_published=True,
        )
        self.foreign_assignment = Assignment.objects.create(
            course=self.other_course,
            title="Homework Physics",
            max_score="100.00",
            is_published=True,
        )

        self.team = self.create_team("Team A", [self.student, self.student2])
        self.foreign_team = self.create_team(
            "Team Foreign", [self.student2], course=self.other_course
        )

    @staticmethod
    def create_user(username: str):
        return User.objects.create_user(
            username=username,
            email=f"{username}@example.com",
            password="pass",
        )

    @staticmethod
    def get_section(course) -> Section:
        """Return the default section of a course, creating it if needed."""
        section, _ = Section.objects.get_or_create(course=course, name="Default")
        return section

    @classmethod
    def enroll(cls, student, course) -> Enrollment:
        return Enrollment.objects.create(
            section=cls.get_section(course),
            student=student,
            status=Status.APPROVED,
        )

    def create_team(self, name, members, course=None, leader=None):
        course = course or self.course
        leader = leader or members[0]
        team = Team.objects.create(
            name=name,
            section=self.get_section(course),
            leader=leader,
        )
        for member in members:
            if member != leader:
                TeamMember.objects.create(team=team, student=member)
        return team

    def authenticate(self, user):
        self.client.force_authenticate(user=user)

    def grade_team(self, assignment, team, score, user=None):
        if user is not None:
            self.authenticate(user)
        return self.client.post(
            reverse(
                "assignment-grade-team",
                kwargs={"assignment_id": assignment.id},
            ),
            {"team": team.id, "score": score},
            format="json",
        )

    def grade_student(self, assignment, student, score, user=None):
        if user is not None:
            self.authenticate(user)
        return self.client.post(
            reverse(
                "assignment-grade-student",
                kwargs={"assignment_id": assignment.id},
            ),
            {"student": student.id, "score": score},
            format="json",
        )


class GradeTeamTests(GradingAPITestCase):
    def test_teacher_can_grade_a_team_of_their_course(self):
        response = self.grade_team(self.assignment, self.team, "95.00", self.teacher)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(Grade.objects.count(), 2)
        self.assertEqual(len(response.data), 2)
        self.assertTrue(
            Grade.objects.filter(
                student=self.student,
                score=Decimal("95.00"),
            ).exists()
        )
        self.assertTrue(
            Grade.objects.filter(
                student=self.student2,
                score=Decimal("95.00"),
            ).exists()
        )

    def test_teacher_cannot_grade_a_team_of_another_course(self):
        response = self.grade_team(self.assignment, self.foreign_team, "95.00", self.teacher)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Grade.objects.count(), 0)

    def test_teacher_cannot_grade_a_course_of_another_teacher(self):
        response = self.grade_team(
            self.foreign_assignment, self.team, "95.00", self.teacher
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(Grade.objects.count(), 0)

    def test_grading_a_team_creates_grades_for_all_members(self):
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)

        member_ids = set(self.team.members.values_list("student_id", flat=True))
        graded_ids = set(Grade.objects.values_list("student_id", flat=True))
        self.assertEqual(graded_ids, member_ids)

    def test_team_grade_marks_all_grades_as_not_individual(self):
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)

        self.assertFalse(
            Grade.objects.filter(
                assignment=self.assignment,
                is_individual=True,
            ).exists()
        )
        self.assertEqual(
            Grade.objects.filter(
                assignment=self.assignment,
                is_individual=False,
            ).count(),
            2,
        )

    def test_grade_team_skips_members_without_approved_enrollment(self):
        """A removed student keeps out of team re-grading (ghost member)."""
        Enrollment.objects.filter(
            student=self.student2, section__course=self.course
        ).delete()

        response = self.grade_team(self.assignment, self.team, "95.00", self.teacher)

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 1)
        self.assertTrue(
            Grade.objects.filter(
                assignment=self.assignment,
                student=self.student,
                score=Decimal("95.00"),
            ).exists()
        )
        self.assertFalse(
            Grade.objects.filter(
                assignment=self.assignment, student=self.student2
            ).exists()
        )

    def test_team_without_active_members_cannot_be_graded(self):
        Enrollment.objects.filter(section__course=self.course).delete()

        response = self.grade_team(self.assignment, self.team, "95.00", self.teacher)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Grade.objects.count(), 0)


class GradeStudentTests(GradingAPITestCase):
    def test_teacher_can_grade_a_student_individually(self):
        response = self.grade_student(
            self.assignment, self.student, "80.00", self.teacher
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        grade = Grade.objects.get(assignment=self.assignment, student=self.student)
        self.assertEqual(grade.score, Decimal("80.00"))
        self.assertTrue(grade.is_individual)
        self.assertEqual(grade.graded_by, self.teacher)

    def test_individual_grade_is_marked_as_individual(self):
        self.grade_student(self.assignment, self.student, "80.00", self.teacher)

        grade = Grade.objects.get(assignment=self.assignment, student=self.student)
        self.assertTrue(grade.is_individual)

    def test_individual_grade_is_not_overwritten_by_team_regrade(self):
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)
        self.grade_student(self.assignment, self.student, "80.00", self.teacher)
        self.grade_team(self.assignment, self.team, "100.00", self.teacher)

        individual = Grade.objects.get(
            assignment=self.assignment, student=self.student
        )
        group_member = Grade.objects.get(
            assignment=self.assignment, student=self.student2
        )
        self.assertEqual(individual.score, Decimal("80.00"))
        self.assertTrue(individual.is_individual)
        self.assertEqual(group_member.score, Decimal("100.00"))
        self.assertFalse(group_member.is_individual)

    def test_team_grade_can_overwrite_individual_grades_when_requested(self):
        """The teacher can force the team score over individual grades."""
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)
        self.grade_student(self.assignment, self.student, "80.00", self.teacher)

        self.authenticate(self.teacher)
        response = self.client.post(
            reverse(
                "assignment-grade-team",
                kwargs={"assignment_id": self.assignment.id},
            ),
            {
                "team": self.team.id,
                "score": "70.00",
                "overwrite_individual": True,
            },
            format="json",
        )

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(len(response.data), 2)
        for student in (self.student, self.student2):
            grade = Grade.objects.get(
                assignment=self.assignment, student=student
            )
            self.assertEqual(grade.score, Decimal("70.00"))
            self.assertFalse(grade.is_individual)

    def test_team_grades_update_when_team_is_regraded(self):
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)
        self.grade_team(self.assignment, self.team, "100.00", self.teacher)

        for student in (self.student, self.student2):
            grade = Grade.objects.get(assignment=self.assignment, student=student)
            self.assertEqual(grade.score, Decimal("100.00"))
            self.assertFalse(grade.is_individual)

    def test_individual_grade_creates_event_log(self):
        response = self.grade_student(
            self.assignment, self.student, "80.00", self.teacher
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        grade = Grade.objects.get(assignment=self.assignment, student=self.student)
        log = EventLog.objects.filter(
            action=EventLog.ACTION_UPDATE, entity_type="grade"
        ).first()
        self.assertIsNotNone(log)
        self.assertEqual(log.actor, self.teacher)
        self.assertEqual(log.target, self.student)
        self.assertEqual(log.entity_id, grade.id)
        self.assertEqual(log.metadata["assignment_id"], self.assignment.id)
        self.assertEqual(log.metadata["score"], "80.00")
        self.assertTrue(log.metadata["is_individual"])

    def test_team_grade_creates_event_log(self):
        response = self.grade_team(self.assignment, self.team, "95.00", self.teacher)
        self.assertEqual(response.status_code, status.HTTP_200_OK)

        log = EventLog.objects.filter(
            action=EventLog.ACTION_UPDATE, entity_type="grade"
        ).first()
        self.assertIsNotNone(log)
        self.assertEqual(log.actor, self.teacher)
        self.assertEqual(log.target, self.team.leader)
        self.assertEqual(log.metadata["assignment_id"], self.assignment.id)
        self.assertEqual(log.metadata["score"], "95.00")
        self.assertEqual(log.metadata["affected"], 2)
        self.assertEqual(log.metadata["team_id"], self.team.id)
        self.assertEqual(log.metadata["team_name"], self.team.name)
        self.assertEqual(
            set(log.metadata["member_ids"]),
            {self.student.id, self.student2.id},
        )


class GradeAccessTests(GradingAPITestCase):
    def test_student_cannot_create_grades(self):
        self.authenticate(self.student)

        team_response = self.grade_team(self.assignment, self.team, "95.00")
        student_response = self.grade_student(
            self.assignment, self.student, "80.00"
        )
        list_response = self.client.post(
            reverse("grade-list"), {}, format="json"
        )

        self.assertEqual(team_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(student_response.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(list_response.status_code, status.HTTP_405_METHOD_NOT_ALLOWED)
        self.assertEqual(Grade.objects.count(), 0)

    def test_student_can_only_list_their_own_grades(self):
        self.grade_student(self.assignment, self.student, "90.00", self.teacher)
        self.grade_student(self.assignment, self.student2, "70.00", self.teacher)

        self.authenticate(self.student)
        response = self.client.get(reverse("grade-list"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["student"]["id"], self.student.id)

    def test_student_cannot_retrieve_another_students_grade(self):
        self.grade_student(self.assignment, self.student2, "70.00", self.teacher)
        grade = Grade.objects.get(
            assignment=self.assignment, student=self.student2
        )

        self.authenticate(self.student)
        response = self.client.get(reverse("grade-detail", args=[grade.id]))

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_teacher_can_list_grades_of_own_courses_only(self):
        self.grade_student(self.assignment, self.student, "90.00", self.teacher)
        self.grade_student(
            self.foreign_assignment, self.student, "60.00", self.other_teacher
        )

        self.authenticate(self.teacher)
        response = self.client.get(reverse("grade-list"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(
            response.data["results"][0]["assignment"]["course"]["id"],
            self.course.id,
        )

    def test_students_do_not_see_grade_origin_but_teachers_do(self):
        """``is_individual`` is teacher-only data."""
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)
        self.grade_student(self.assignment, self.student, "80.00", self.teacher)

        self.authenticate(self.student)
        student_response = self.client.get(reverse("grade-list"))

        self.assertEqual(student_response.status_code, status.HTTP_200_OK)
        self.assertGreater(student_response.data["count"], 0)
        for item in student_response.data["results"]:
            self.assertNotIn("is_individual", item)

        self.authenticate(self.teacher)
        teacher_response = self.client.get(reverse("grade-list"))

        self.assertEqual(teacher_response.status_code, status.HTTP_200_OK)
        for item in teacher_response.data["results"]:
            self.assertIn("is_individual", item)

    def test_student_loses_access_to_grades_after_enrollment_removal(self):
        """Grades are preserved but hidden from the removed student."""
        self.grade_student(self.assignment, self.student, "90.00", self.teacher)

        Enrollment.objects.filter(
            student=self.student, section__course=self.course
        ).delete()

        self.authenticate(self.student)
        student_response = self.client.get(reverse("grade-list"))
        detail = Grade.objects.get(assignment=self.assignment, student=self.student)
        detail_response = self.client.get(
            reverse("grade-detail", args=[detail.id])
        )

        self.assertEqual(student_response.status_code, status.HTTP_200_OK)
        self.assertEqual(student_response.data["count"], 0)
        self.assertEqual(detail_response.status_code, status.HTTP_404_NOT_FOUND)

        # The teacher keeps the historical record of their course.
        self.authenticate(self.teacher)
        teacher_response = self.client.get(reverse("grade-list"))

        self.assertEqual(teacher_response.status_code, status.HTTP_200_OK)
        self.assertEqual(teacher_response.data["count"], 1)


class GradeValidationTests(GradingAPITestCase):
    def test_cannot_grade_student_not_belonging_to_course(self):
        response = self.grade_student(
            self.assignment, self.unapproved_student, "50.00", self.teacher
        )

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Grade.objects.count(), 0)

    def test_cannot_grade_a_team_belonging_to_another_course(self):
        response = self.grade_team(self.assignment, self.foreign_team, "95.00", self.teacher)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Grade.objects.count(), 0)

    def test_cannot_exceed_assignment_max_score(self):
        team_response = self.grade_team(self.assignment, self.team, "101.00", self.teacher)
        student_response = self.grade_student(
            self.assignment, self.student, "101.00", self.teacher
        )

        self.assertEqual(team_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(student_response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Grade.objects.count(), 0)

    def test_cannot_use_a_negative_score(self):
        response = self.grade_team(self.assignment, self.team, "-5.00", self.teacher)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Grade.objects.count(), 0)

    def test_cannot_grade_a_team_without_members(self):
        empty_team = self.create_team("Empty Team", [self.unapproved_student])
        TeamMember.objects.filter(team=empty_team).delete()

        response = self.grade_team(self.assignment, empty_team, "95.00", self.teacher)

        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Grade.objects.count(), 0)

    def test_only_one_grade_per_assignment_and_student(self):
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)
        self.grade_team(self.assignment, self.team, "100.00", self.teacher)

        self.assertEqual(
            Grade.objects.filter(
                assignment=self.assignment, student=self.student
            ).count(),
            1,
        )
        grade = Grade.objects.get(
            assignment=self.assignment, student=self.student
        )
        self.assertEqual(grade.score, Decimal("100.00"))

        with self.assertRaises(IntegrityError):
            Grade.objects.create(
                assignment=self.assignment,
                student=self.student,
                score="90.00",
                graded_by=self.teacher,
            )

    def test_failed_grading_does_not_create_event_log(self):
        response = self.grade_student(
            self.assignment, self.unapproved_student, "50.00", self.teacher
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(
            EventLog.objects.filter(
                action=EventLog.ACTION_UPDATE, entity_type="grade"
            ).exists()
        )
        response = self.grade_team(
            self.assignment, self.foreign_team, "95.00", self.teacher
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(
            EventLog.objects.filter(
                action=EventLog.ACTION_UPDATE, entity_type="grade"
            ).exists()
        )


class SectionGradesExportTests(GradingAPITestCase):
    """GET /api/sections/{id}/export-grades/ (Excel workbook download)."""

    def setUp(self):
        super().setUp()
        self.section = self.get_section(self.course)
        self.ungraded_student = self.create_user("ungraded_student")
        self.enroll(self.ungraded_student, self.course)

    def export_url(self) -> str:
        return reverse(
            "section-export-grades", kwargs={"pk": self.section.id}
        )

    def test_teacher_downloads_workbook_with_expected_layout(self):
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)
        self.grade_student(self.assignment, self.student2, "80.00", self.teacher)
        Assignment.objects.create(
            course=self.course,
            title="Draft homework",
            max_score="50.00",
            is_published=False,
        )
        self.enroll(self.unapproved_student, self.course)
        Enrollment.objects.filter(student=self.unapproved_student).update(
            status=Status.PENDING
        )

        self.authenticate(self.teacher)
        response = self.client.get(self.export_url())

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("attachment", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        self.assertEqual(sheet["A1"].value, "Curso:")
        self.assertEqual(sheet["B1"].value, "Math 101")
        self.assertEqual(sheet["A2"].value, "Grupo:")
        self.assertEqual(sheet["B2"].value, "Default")
        self.assertEqual(
            [cell.value for cell in sheet[4]],
            ["Estudiante", "Homework 1", "Total"],
        )
        rows = {row[0].value: row for row in sheet.iter_rows(min_row=5)}
        student_row = rows[self.student.username]
        student2_row = rows[self.student2.username]
        self.assertEqual(student_row[1].value, 95.0)
        self.assertEqual(student_row[2].value, 95.0)
        self.assertEqual(student2_row[1].value, 80.0)
        self.assertEqual(student2_row[2].value, 80.0)
        self.assertNotIn(self.unapproved_student.username, rows)

    def test_total_sums_only_existing_grades(self):
        self.grade_team(self.assignment, self.team, "95.00", self.teacher)

        self.authenticate(self.teacher)
        response = self.client.get(self.export_url())

        workbook = load_workbook(BytesIO(response.content))
        sheet = workbook.active
        rows = {row[0].value: row for row in sheet.iter_rows(min_row=5)}
        ungraded_row = rows[self.ungraded_student.username]
        self.assertIsNone(ungraded_row[1].value)
        self.assertEqual(ungraded_row[2].value, 0)

    def test_other_teacher_cannot_export(self):
        self.authenticate(self.other_teacher)
        response = self.client.get(self.export_url())

        self.assertEqual(response.status_code, status.HTTP_404_NOT_FOUND)

    def test_student_cannot_export(self):
        self.authenticate(self.student)
        response = self.client.get(self.export_url())

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

class SuperuserIsolationTests(GradingAPITestCase):
    """The superuser has no special powers in the regular grading views."""

    def setUp(self):
        super().setUp()
        self.superuser = self.create_user("root")
        self.superuser.is_superuser = True
        self.superuser.save()

    def test_superuser_sees_no_grades(self):
        Grade.objects.create(
            assignment=self.assignment,
            student=self.student,
            score=Decimal("95.00"),
            graded_by=self.teacher,
        )
        self.authenticate(self.superuser)

        response = self.client.get("/api/grades/")

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.data["count"], 0)

    def test_superuser_cannot_grade_foreign_assignment(self):
        response = self.grade_student(
            self.foreign_assignment,
            self.student2,
            "50.00",
            self.superuser,
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_superuser_cannot_grade_team_of_foreign_course(self):
        response = self.grade_team(
            self.foreign_assignment,
            self.foreign_team,
            "50.00",
            self.superuser,
        )

        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)
